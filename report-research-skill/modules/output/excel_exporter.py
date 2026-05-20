"""
Excel 导出模块
将报告数据导出为结构化 Excel 文件
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional
from datetime import datetime


# 样式定义
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_workbook() -> openpyxl.Workbook:
    """创建新的 Excel 工作簿"""
    return openpyxl.Workbook()


def style_header(cell):
    """为表头单元格应用样式"""
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CELL_ALIGNMENT
    cell.border = THIN_BORDER


def style_data_cell(cell):
    """为数据单元格应用样式"""
    cell.alignment = CELL_ALIGNMENT
    cell.border = THIN_BORDER


def export_metadata_sheet(wb: openpyxl.Workbook, reports: List[Dict]):
    """导出报告元数据 Sheet"""
    ws = wb.active
    ws.title = "报告元数据"

    # 表头
    headers = ["标题", "来源", "发布日期", "作者/机构", "链接", "可靠性评级"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        style_header(cell)

    # 数据
    for row, report in enumerate(reports, 2):
        ws.cell(row=row, column=1).value = report.get('title', 'N/A')
        ws.cell(row=row, column=2).value = report.get('source', 'N/A')
        ws.cell(row=row, column=3).value = report.get('publish_date', 'N/A')
        ws.cell(row=row, column=4).value = report.get('authors', 'N/A')
        ws.cell(row=row, column=5).value = report.get('url', 'N/A')
        ws.cell(row=row, column=6).value = report.get('confidence_level', 'N/A')

        # 应用样式
        for col in range(1, 7):
            style_data_cell(ws.cell(row=row, column=col))

    # 调整列宽
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 12


def export_comparison_sheet(wb: openpyxl.Workbook, reports: List[Dict]):
    """导出结论对比 Sheet"""
    ws = wb.create_sheet("结论对比")

    # 找出最大结论数
    max_findings = max(
        len(r.get('key_findings', [])) if r.get('key_findings') else 0
        for r in reports
    )

    # 表头
    headers = ["标题", "来源", "日期"]
    for i in range(1, max_findings + 1):
        headers.append(f"核心结论 {i}")
    headers.extend(["研究方法/数据来源", "市场展望", "局限性"])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        style_header(cell)

    # 数据
    for row, report in enumerate(reports, 2):
        col = 1
        ws.cell(row=row, column=col).value = report.get('title', 'N/A')
        style_data_cell(ws.cell(row=row, column=col))
        col += 1

        ws.cell(row=row, column=col).value = report.get('source', 'N/A')
        style_data_cell(ws.cell(row=row, column=col))
        col += 1

        ws.cell(row=row, column=col).value = report.get('publish_date', 'N/A')
        style_data_cell(ws.cell(row=row, column=col))
        col += 1

        # 核心结论
        findings = report.get('key_findings', [])
        for i in range(max_findings):
            if i < len(findings):
                ws.cell(row=row, column=col).value = findings[i]
            else:
                ws.cell(row=row, column=col).value = "未提及"
            style_data_cell(ws.cell(row=row, column=col))
            col += 1

        # 研究方法
        ws.cell(row=row, column=col).value = report.get('methodology', 'N/A')
        style_data_cell(ws.cell(row=row, column=col))
        col += 1

        # 市场展望
        ws.cell(row=row, column=col).value = report.get('market_outlook', 'N/A')
        style_data_cell(ws.cell(row=row, column=col))
        col += 1

        # 局限性
        ws.cell(row=row, column=col).value = report.get('limitations', 'N/A')
        style_data_cell(ws.cell(row=row, column=col))

    # 调整列宽
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    for i in range(4, 4 + max_findings):
        ws.column_dimensions[get_column_letter(i)].width = 40
    ws.column_dimensions[get_column_letter(4 + max_findings)].width = 25
    ws.column_dimensions[get_column_letter(5 + max_findings)].width = 30
    ws.column_dimensions[get_column_letter(6 + max_findings)].width = 30


def export_statistics_sheet(wb: openpyxl.Workbook, reports: List[Dict], topic: str):
    """导出统计 Sheet"""
    ws = wb.create_sheet("统计分析")

    # 来源统计
    ws['A1'] = "统计分析"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')

    row = 3
    ws.cell(row=row, column=1).value = "来源分布"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)

    row += 1
    ws.cell(row=row, column=1).value = "来源"
    ws.cell(row=row, column=2).value = "数量"
    ws.cell(row=row, column=3).value = "占比"
    for col in range(1, 4):
        style_header(ws.cell(row=row, column=col))

    source_counts = {}
    for report in reports:
        source = report.get('source', 'Unknown')
        source_counts[source] = source_counts.get(source, 0) + 1

    total = len(reports)
    for source, count in sorted(source_counts.items(), key=lambda x: -x[1]):
        row += 1
        ws.cell(row=row, column=1).value = source
        ws.cell(row=row, column=2).value = count
        ws.cell(row=row, column=3).value = f"{count / total * 100:.1f}%"
        for col in range(1, 4):
            style_data_cell(ws.cell(row=row, column=col))

    # 可靠性统计
    row += 2
    ws.cell(row=row, column=1).value = "可靠性分布"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)

    row += 1
    ws.cell(row=row, column=1).value = "评级"
    ws.cell(row=row, column=2).value = "数量"
    ws.cell(row=row, column=3).value = "占比"
    for col in range(1, 4):
        style_header(ws.cell(row=row, column=col))

    confidence_counts = {'高': 0, '中': 0, '低': 0}
    for report in reports:
        level = report.get('confidence_level', '中')
        if level in confidence_counts:
            confidence_counts[level] += 1

    for level in ['高', '中', '低']:
        row += 1
        count = confidence_counts[level]
        ws.cell(row=row, column=1).value = level
        ws.cell(row=row, column=2).value = count
        ws.cell(row=row, column=3).value = f"{count / total * 100:.1f}%"
        for col in range(1, 4):
            style_data_cell(ws.cell(row=row, column=col))

    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def export_to_excel(reports: List[Dict], topic: str, output_path: Optional[str] = None) -> str:
    """
    将报告数据导出为 Excel 文件

    Args:
        reports: 报告字典列表
        topic: 研究主题
        output_path: 输出文件路径（可选，默认使用主题名）

    Returns:
        导出的文件路径
    """
    if output_path is None:
        # 生成默认文件名
        safe_topic = "".join(c if c.isalnum() or c in ('-', '_') else '_' for c in topic)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"{safe_topic}_report_analysis_{timestamp}.xlsx"

    wb = create_workbook()

    # 导出各个 Sheet
    export_metadata_sheet(wb, reports)
    export_comparison_sheet(wb, reports)
    export_statistics_sheet(wb, reports, topic)

    # 保存
    wb.save(output_path)

    return output_path


# 示例用法
if __name__ == '__main__':
    sample_reports = [
        {
            'title': 'Attention Is All You Need',
            'source': 'ArXiv',
            'publish_date': '2017-06-12',
            'authors': 'Vaswani et al.',
            'url': 'https://arxiv.org/abs/1706.03762',
            'confidence_level': '高',
            'key_findings': [
                'Transformer achieves state-of-the-art results',
                'Training time reduced by 60%',
                'BLEU score 28.4 on WMT 2014'
            ],
            'methodology': 'Neural Network Architecture',
            'market_outlook': 'Foundational for modern NLP',
            'limitations': 'Computationally expensive'
        },
        {
            'title': 'GPT-4 Technical Report',
            'source': 'OpenAI',
            'publish_date': '2023-03-15',
            'authors': 'OpenAI Team',
            'url': 'https://arxiv.org/abs/2303.08774',
            'confidence_level': '高',
            'key_findings': [
                'Human-level performance on professional exams',
                'Multimodal capabilities',
                'Reduced hallucinations compared to GPT-3.5'
            ],
            'methodology': 'Large Language Model Pre-training',
            'market_outlook': 'Enterprise AI adoption accelerating',
            'limitations': 'API-only access, closed source'
        }
    ]

    output_file = export_to_excel(sample_reports, "Large Language Models")
    print(f"Excel exported to: {output_file}")
